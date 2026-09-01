"""
Excel -> SQLite SQL 脚本生成器
用法:
    python excel_to_sqlite.py <excel文件路径> [--sheet 工作表名] [--table 表名] [--output 输出.sql]

示例:
    python excel_to_sqlite.py C:/data/orders.xlsx --table orders --output orders.sql

说明:
    - 读取 Excel 第一个工作表(或 --sheet 指定)
    - 第一行为列名
    - 根据每列内容自动推断 SQLite 字段类型 (INTEGER/REAL/TEXT)
    - 输出包含 CREATE TABLE + INSERT INTO 的 .sql 文件
    - 所有标识符(表名/列名)自动加双引号, 字符串自动转义单引号
"""

import sys
import os
import re
import argparse
from datetime import datetime, date, time

try:
    import openpyxl
except ImportError:
    print("缺少依赖 openpyxl, 请先安装:  pip install openpyxl")
    sys.exit(1)


def infer_type(values):
    """根据一列的值推断 SQLite 类型: INTEGER / REAL / TEXT"""
    has_text = False
    has_real = False
    has_int = False
    for v in values:
        if v is None:
            continue
        if isinstance(v, bool):
            has_int = True
            continue
        if isinstance(v, int):
            has_int = True
            continue
        if isinstance(v, float):
            has_real = True
            continue
        if isinstance(v, (datetime, date, time)):
            has_text = True
            continue
        # 字符串
        s = str(v).strip()
        if s == "":
            continue
        # 尝试转数字
        if re.fullmatch(r"[-+]?\d+", s):
            has_int = True
            continue
        if re.fullmatch(r"[-+]?\d+\.\d+", s) or re.fullmatch(r"[-+]?\d*\.\d+([eE][-+]?\d+)?", s):
            has_real = True
            continue
        has_text = True
    if has_text:
        return "TEXT"
    if has_real:
        return "REAL"
    if has_int:
        return "INTEGER"
    return "TEXT"


def clean_identifier(name):
    """列名/表名清洗: 去空白, 非法字符替换为下划线"""
    name = str(name).strip()
    name = re.sub(r'[^\w\u4e00-\u9fff]', '_', name)
    return name or "col"


# 数据中代表空值的字面量（如 MySQL 导出的 '\N'），统一转成 SQL NULL
NULL_LITERALS = {""}

def format_value(v, sql_type):
    """格式化单值用于 INSERT 语句"""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, (datetime, date, time)):
        return "'" + v.isoformat() + "'"
    # 强制 TEXT 类型时统一按字符串输出（保留前导零、数字也加引号）
    if sql_type.upper() == "TEXT":
        s = str(v).strip()
        if s in NULL_LITERALS or s == "\\N":
            return "NULL"
        s = s.replace("'", "''")
        return "'" + s + "'"
    # 非 TEXT：数字原样, 字符串/其它按原样
    if isinstance(v, (int, float)):
        return str(v) if isinstance(v, int) else repr(v)
    s = str(v).strip()
    if s in NULL_LITERALS or s == "\\N":
        return "NULL"
    s = s.replace("'", "''")
    return "'" + s + "'"


def parse_col_types(spec):
    """解析 --col-types 'name:TEXT,adcode:TEXT' 为 {列名(小写): 类型}"""
    result = {}
    if not spec:
        return result
    for part in spec.split(","):
        part = part.strip()
        if not part or ":" not in part:
            continue
        col, typ = part.rsplit(":", 1)
        result[col.strip().lower()] = typ.strip().upper()
    return result


def main():
    parser = argparse.ArgumentParser(description="Excel 导出为 SQLite SQL 脚本")
    parser.add_argument("excel", help="Excel 文件路径 (.xlsx)")
    parser.add_argument("--sheet", default=None, help="工作表名, 默认第一个")
    parser.add_argument("--table", default=None, help="目标表名, 默认用 Excel 文件名")
    parser.add_argument("--output", default=None, help="输出 .sql 文件路径, 默认同目录同名 .sql")
    parser.add_argument("--col-types", default=None,
                        help="强制指定列类型, 逗号分隔, 如 'name:TEXT,adcode:TEXT,citycode:TEXT'")
    parser.add_argument("--no-create-table", action="store_true",
                        help="只输出 INSERT 语句, 不生成 CREATE TABLE(表结构由外部定义)")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"文件不存在: {args.excel}")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows or not any(r is not None for r in rows[0]):
        print("Excel 为空或没有表头")
        sys.exit(1)

    # 列名: 第一行, 过滤全空列
    headers = []
    for i, cell in enumerate(rows[0]):
        if cell is None or str(cell).strip() == "":
            continue
        headers.append((i, clean_identifier(cell)))

    # 提取各列数据
    col_values = {i: [] for i, _ in headers}
    for row in rows[1:]:
        for i, _ in headers:
            col_values[i].append(row[i] if i < len(row) else None)

    # 推断类型, 可用 --col-types 强制覆盖
    types = {i: infer_type(vals) for i, vals in col_values.items()}
    forced = parse_col_types(args.col_types)
    for i, name in headers:
        if name.lower() in forced:
            types[i] = forced[name.lower()]

    table = clean_identifier(args.table) if args.table else clean_identifier(os.path.splitext(os.path.basename(args.excel))[0])
    output = args.output or os.path.join(os.path.dirname(os.path.abspath(args.excel)),
                                         os.path.splitext(os.path.basename(args.excel))[0] + ".sql")

    sql = []
    if not args.no_create_table:
        col_defs = [f'  "{name}" {types[i]}' for i, name in headers]
        sql.append(f'CREATE TABLE IF NOT EXISTS "{table}" (')
        sql.append(",\n".join(col_defs))
        sql.append(");\n")

    # 生成 INSERT
    insert_cols = ", ".join(f'"{name}"' for _, name in headers)
    sql.append(f"INSERT INTO \"{table}\" ({insert_cols}) VALUES")
    values_list = []
    for idx, row in enumerate(rows[1:]):
        if all(row[i] is None for i, _ in headers):
            continue
        vals = [format_value(row[i], types[i]) if i < len(row) else "NULL" for i, _ in headers]
        values_list.append("(" + ", ".join(vals) + ")")
    if values_list:
        sql.append(",\n".join(values_list) + ";")
    else:
        sql.append("(NULL);")  # 无数据时给个空占位, 避免非法 SQL

    with open(output, "w", encoding="utf-8") as f:
        f.write("\n".join(sql) + "\n")

    print(f"已生成: {output}")
    print(f"表名: {table}, 列: {len(headers)} 个, 数据行: {len(values_list)} 行")
    print("字段定义:")
    for i, name in headers:
        print(f"  {name}  {types[i]}")


if __name__ == "__main__":
    main()
